from __future__ import annotations

import argparse
from decimal import Decimal, InvalidOperation
from pathlib import Path

from sqlalchemy import func, select

from apps.api.app.db.session import SessionLocal
from apps.api.app.models.campaign import CampaignProduct
from apps.api.app.models.product import Product


def decimal_or_none(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", ".")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def norm(value: str | None) -> str:
    return " ".join((value or "").strip().lower().split())


def detect_header(row: list) -> bool:
    labels = [norm(str(cell)) for cell in row if cell is not None]
    joined = " | ".join(labels)
    return (
        "produtor" in joined
        and ("preço venda" in joined or "preco venda" in joined)
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Import Alice spreadsheet pricing into products/campaign_products. "
            "Expected columns: description, producer, preço revenda, preço venda."
        ),
    )
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--campaign-id", default=None)
    args = parser.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for this script. Install it in the API image "
            "or run: pip install openpyxl"
        ) from exc

    wb = load_workbook(args.xlsx_path, data_only=True)
    db = SessionLocal()

    products_updated = 0
    campaign_products_updated = 0
    unmatched: list[str] = []

    try:
        for sheet in wb.worksheets:
            header_map: dict[str, int] | None = None

            for row in sheet.iter_rows(values_only=True):
                values = list(row)

                if detect_header(values):
                    header_map = {}
                    for idx, cell in enumerate(values):
                        label = norm(str(cell))
                        if "descr" in label or "item" in label or "produto" in label:
                            header_map.setdefault("name", idx)
                        elif "produtor" in label:
                            header_map["producer"] = idx
                        elif "revenda" in label:
                            header_map["cost"] = idx
                        elif "venda" in label:
                            header_map["sale"] = idx
                    continue

                if not header_map:
                    continue

                name = values[header_map.get("name", 0)]
                if not name:
                    continue

                name_s = str(name).strip()
                if name_s.upper() == name_s and len(name_s) > 3:
                    # Section label such as MERCEARIA or HORTIFRUTI.
                    continue

                producer = (
                    str(values[header_map["producer"]]).strip()
                    if "producer" in header_map and values[header_map["producer"]]
                    else None
                )
                cost_price = (
                    decimal_or_none(values[header_map["cost"]])
                    if "cost" in header_map
                    else None
                )
                sale_price = (
                    decimal_or_none(values[header_map["sale"]])
                    if "sale" in header_map
                    else None
                )

                if cost_price is None and sale_price is None:
                    continue

                product = db.scalars(
                    select(Product).where(
                        func.lower(Product.name) == norm(name_s),
                    ),
                ).first()

                if product is None:
                    unmatched.append(name_s)
                else:
                    product.producer_name_snapshot = (
                        producer or product.producer_name_snapshot
                    )
                    product.cost_price = cost_price
                    product.sale_price = sale_price or product.price
                    product.price = product.sale_price
                    products_updated += 1

                stmt = select(CampaignProduct).where(
                    func.lower(CampaignProduct.name_snapshot) == norm(name_s),
                )
                if args.campaign_id:
                    stmt = stmt.where(CampaignProduct.campaign_id == args.campaign_id)

                campaign_products = db.scalars(stmt).all()
                for campaign_product in campaign_products:
                    sale = sale_price or campaign_product.price
                    cost = cost_price or Decimal("0.00")
                    campaign_product.producer_name_snapshot = (
                        producer or campaign_product.producer_name_snapshot
                    )
                    campaign_product.cost_price_snapshot = cost
                    campaign_product.sale_price_snapshot = sale
                    campaign_product.margin_unit_snapshot = sale - cost
                    campaign_product.price = sale
                    campaign_products_updated += 1

        db.commit()
        print(f"Products updated: {products_updated}")
        print(f"Campaign products updated: {campaign_products_updated}")
        print(f"Unmatched spreadsheet products: {len(set(unmatched))}")
        for item in sorted(set(unmatched))[:100]:
            print(f"- {item}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
